#!/usr/bin/env python3
"""
Unified Portal seed generator — builds the org tree from the REAL Gujarat
school registry (Docs/Guj School Master.xlsx). IDs nest exactly
(District 4 -> Block 6 -> Cluster 10 -> School 11). Real district/block/
cluster/school NAMES + management/medium attributes are kept; grades are
generated from each school's lowclass..highclass and 1-3 sections per grade.
People (teachers/principals/students) are ANONYMISED by default (privacy);
flip USE_REAL_PRINCIPAL to surface the registry's NamePrincipal.

Run: `npm run seed` (-> python scripts/generateSeed.py).
"""
import json, os, sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "..", "Docs", "Guj School Master.xlsx")
OUT = os.path.join(ROOT, "src", "data", "seed")
SUPA = os.path.join(ROOT, "supabase")

# representative subset (large + populated) — ~10 districts / ~1,000 schools
N_DISTRICTS, N_BLOCKS, N_CLUSTERS, N_SCHOOLS = 10, 4, 5, 5
USE_REAL_PRINCIPAL = False  # anonymise principals by default (one-line switch)
STATE_CODE = "24"

FIRST = ["Anjali","Ramesh","Priya","Mahesh","Nita","Vikram","Geeta","Sunil","Kiran","Hetal",
         "Jignesh","Falguni","Bhavna","Rohit","Sneha","Amit","Komal","Dilip","Reena","Paresh",
         "Manisha","Tushar","Asha","Nilesh","Devang","Roshni","Harshad","Ila","Mitesh","Vaishali"]
LAST = ["Patel","Solanki","Joshi","Rabari","Chauhan","Desai","Parmar","Vasava","Makwana","Bhatt",
        "Modi","Shah","Thakor","Mehta","Pandya","Dave","Trivedi","Gohil","Rathod","Vyas"]

def h(s):
    x = 2166136261
    for c in str(s):
        x ^= ord(c); x = (x * 16777619) & 0xFFFFFFFF
    return x

def person(seed): return f"{FIRST[h(seed+'a')%len(FIRST)]} {LAST[h(seed+'b')%len(LAST)]}"
def title(s): return " ".join(w.capitalize() for w in str(s).strip().split())
def jit(seed, spread): return ((h(seed) % 1000)/1000 - 0.5) * 2 * spread

print("Reading registry…", file=sys.stderr)
wb = load_workbook(XLSX, data_only=True, read_only=True)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True)
hdr = [str(c) for c in next(it)]
col = {name: i for i, name in enumerate(hdr)}

# group: district -> block -> cluster -> [school rows]
tree = {}
for r in it:
    did = r[col["DistrictId"]]
    if did is None: continue
    did = str(did).strip(); bid = str(r[col["BlockId"]]).strip(); cid = str(r[col["ClusterId"]]).strip(); sid = str(r[col["SchoolId"]]).strip()
    if not (did.isdigit() and bid.isdigit() and cid.isdigit() and sid.isdigit()): continue
    rec = {
        "sid": sid, "did": did, "bid": bid, "cid": cid,
        "district": r[col["District"]], "block": r[col["Block"]], "cluster": r[col["Cluster"]],
        "school": r[col["School"]], "mgt": r[col["SchMgt_desc"]],
        "low": r[col["lowclass"]], "high": r[col["highclass"]],
        "principal": r[col["NamePrincipal"]], "medium": r[col["Medium1_desc"]],
    }
    tree.setdefault(did, {}).setdefault(bid, {}).setdefault(cid, []).append(rec)

# pick districts that have the required nested shape
def shaped_blocks(blocks):
    return [b for b, cl in blocks.items() if sum(1 for s in cl.values() if len(s) >= N_SCHOOLS) >= N_CLUSTERS]
cands = [(did, b) for did, b in tree.items() if len(shaped_blocks(b)) >= N_BLOCKS]
cands.sort(key=lambda x: x[0])
chosen = cands[:N_DISTRICTS]
if len(chosen) < N_DISTRICTS:
    sys.exit("Not enough districts with the required shape.")

entities = []; app_users = []; demo = []
state = {"id": "st-gj", "name": "Gujarat", "name_gu": "ગુજરાત", "level": "state", "parent_id": None, "meta": {"code": STATE_CODE}}
entities.append(state)
teacher_seq = 1; principal_seq = 1
sections_for_users = []

for did, blocks in chosen:
    drec0 = next(iter(next(iter(blocks.values())).values()))[0]
    dId = f"dist-{did}"
    entities.append({"id": dId, "name": title(drec0["district"]), "name_gu": title(drec0["district"]),
                     "level": "district", "parent_id": state["id"], "meta": {"code": did}})
    for bid, clusters in [(b, blocks[b]) for b in shaped_blocks(blocks)[:N_BLOCKS]]:
        brec0 = next(iter(clusters.values()))[0]
        bId = f"block-{bid}"
        entities.append({"id": bId, "name": title(brec0["block"]), "name_gu": title(brec0["block"]),
                         "level": "block", "parent_id": dId, "meta": {"code": bid}})
        good_clusters = [(c, s) for c, s in clusters.items() if len(s) >= N_SCHOOLS][:N_CLUSTERS]
        for cid, schools in good_clusters:
            cId = f"cluster-{cid}"
            entities.append({"id": cId, "name": title(schools[0]["cluster"]), "name_gu": title(schools[0]["cluster"]),
                             "level": "cluster", "parent_id": bId, "meta": {"code": cid, "total_schools": N_SCHOOLS}})
            for sc in schools[:N_SCHOOLS]:
                sid = sc["sid"]; sId = f"sch-{sid}"
                anchor = round(0.45 + (h(sid) % 1000)/1000 * 0.45, 3)  # 0.45–0.90
                try: low, high = int(sc["low"]), int(sc["high"])
                except (TypeError, ValueError): low, high = 1, 8
                if low < 1 or high < low or high > 12: low, high = 1, 8
                enrol = 130 + (h(sid) % 420); teachers = 5 + (h(sid+"t") % 16)
                entities.append({
                    "id": sId, "name": title(sc["school"]), "name_gu": title(sc["school"]),
                    "level": "school", "parent_id": cId,
                    "meta": {
                        "code": sid, "udise_code": sid, "anchor": anchor,
                        "pmShri": h(sid+"pm") % 6 == 0, "enrolment": enrol, "teachers": teachers,
                        "lowClass": low, "highClass": high,
                        "schMgt": sc["mgt"], "medium": (str(sc["medium"]).split("-")[-1] if sc["medium"] else None),
                        "principalName": (title(sc["principal"]) if USE_REAL_PRINCIPAL and sc["principal"] else person("prin"+sId)),
                    },
                })
                # grades from the school's real class range; 1-3 sections each
                for g in range(low, high + 1):
                    gId = f"{sId}-g{g}"
                    gAnchor = round(min(1, max(0, anchor + jit(gId, 0.08))), 3)
                    entities.append({"id": gId, "name": f"Grade {g}", "name_gu": f"ધોરણ {g}",
                                     "level": "grade", "parent_id": sId, "meta": {"grade_no": g, "anchor": gAnchor}})
                    n_sec = [2, 2, 1, 3][h(gId) % 4]
                    for si in range(n_sec):
                        sec = chr(ord("A") + si); secId = f"{gId}-{sec}"
                        tcode = STATE_CODE + str(teacher_seq).zfill(8)  # 10-digit teacher id
                        teacher_seq += 1
                        secAnchor = round(min(1, max(0, gAnchor + jit(secId, 0.07))), 3)
                        students = 18 + (h(secId) % 20)
                        entities.append({"id": secId, "name": f"{g}-{sec}", "name_gu": f"{g}-{sec}",
                                         "level": "section", "parent_id": gId,
                                         "meta": {"grade_no": g, "section_label": sec, "teacher_name": person(secId),
                                                  "teacher_id": tcode, "students": students, "anchor": secAnchor}})
                        sections_for_users.append((secId, sId, sid, tcode))

# recompute non-leaf anchors as mean of school descendants
by_id = {e["id"]: e for e in entities}
kids = {}
for e in entities:
    if e["parent_id"]: kids.setdefault(e["parent_id"], []).append(e["id"])
def school_anchors(eid):
    e = by_id[eid]
    if e["level"] in ("school", "grade", "section"): return [e["meta"]["anchor"]]
    out = []
    for k in kids.get(eid, []): out += school_anchors(k)
    return out
for e in entities:
    if e["level"] in ("state", "district", "block", "cluster"):
        a = school_anchors(e["id"]); e["meta"]["anchor"] = round(sum(a)/len(a), 3) if a else 0.6
        if e["level"] != "state": e["meta"]["total_schools"] = len(a)
state["meta"]["total_schools"] = sum(1 for e in entities if e["level"] == "school")

# app_users — Teacher/Principal use 10-digit ID + 11-digit School UDISE; officers use code + 4-digit PIN
d0 = next(e for e in entities if e["level"] == "district")
b0 = next(e for e in entities if e["level"] == "block" and e["parent_id"] == d0["id"])
c0 = next(e for e in entities if e["level"] == "cluster" and e["parent_id"] == b0["id"])
s0 = next(e for e in entities if e["level"] == "school" and e["parent_id"] == c0["id"])
sec0 = next(e for e in entities if e["level"] == "section" and e["id"].startswith(s0["id"] + "-g"))
prin_id = STATE_CODE + str(900000001 + principal_seq)[1:]  # 10-digit principal id

def add(u): app_users.append(u); demo.append(u)
add({"id": "u-teacher", "login_id": sec0["meta"]["teacher_id"], "name": sec0["meta"]["teacher_name"], "name_gu": "",
     "role": "teacher", "designation": "Teacher", "entity_id": sec0["id"], "school_id": s0["meta"]["code"], "active": True})
add({"id": "u-principal", "login_id": prin_id, "name": person("prin0"), "name_gu": "",
     "role": "principal", "designation": "Principal", "entity_id": s0["id"], "school_id": s0["meta"]["code"], "active": True})
add({"id": "u-crc", "login_id": c0["meta"]["code"], "name": person("crc"), "name_gu": "", "role": "crc",
     "designation": "Cluster Resource Coordinator", "entity_id": c0["id"], "passcode": "1234", "active": True})
add({"id": "u-brc", "login_id": b0["meta"]["code"], "name": person("brc"), "name_gu": "", "role": "brc",
     "designation": "Block Resource Coordinator (BEO)", "entity_id": b0["id"], "passcode": "2345", "active": True})
add({"id": "u-deo", "login_id": d0["meta"]["code"], "name": person("deo"), "name_gu": "", "role": "deo",
     "designation": "District Education Officer", "entity_id": d0["id"], "passcode": "3456", "active": True})
add({"id": "u-state", "login_id": STATE_CODE, "name": "State VSK Cell", "name_gu": "રાજ્ય VSK કક્ષ", "role": "state",
     "designation": "State Project Director", "entity_id": "st-gj", "passcode": "0000", "active": True})
# extra teachers/principals so leaderboards have names
for i, (secid, sid_e, sUd, tcode) in enumerate([x for x in sections_for_users if "-g7-" in x[0]][:4]):
    app_users.append({"id": f"u-teacher-{i}", "login_id": by_id[secid]["meta"]["teacher_id"],
                      "name": by_id[secid]["meta"]["teacher_name"], "name_gu": "", "role": "teacher",
                      "designation": "Teacher", "entity_id": secid, "school_id": sUd, "active": True})

os.makedirs(OUT, exist_ok=True); os.makedirs(SUPA, exist_ok=True)
json.dump(entities, open(os.path.join(OUT, "entities.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=2)
json.dump(app_users, open(os.path.join(OUT, "appUsers.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=2)
counts = {}
for e in entities: counts[e["level"]] = counts.get(e["level"], 0) + 1
counts["total"] = len(entities)
meta = {
    "generatedFrom": "Docs/Guj School Master.xlsx (real Gujarat registry, sliced subset)",
    "idFormats": {"state": 2, "district": 4, "block": 6, "cluster": 10, "school": 11, "teacherPrincipalId": 10, "pin": 4},
    "principalsAnonymised": not USE_REAL_PRINCIPAL,
    "shape": {"districts": N_DISTRICTS, "blocksPerDistrict": N_BLOCKS, "clustersPerBlock": N_CLUSTERS, "schoolsPerCluster": N_SCHOOLS},
    "counts": counts,
    "districts": [title(next(iter(next(iter(b.values())).values()))[0]["district"]) for _, b in chosen],
    "pmShriSchools": sum(1 for e in entities if e["level"] == "school" and e["meta"].get("pmShri")),
    "demoLogins": [{"role": u["role"], "login_id": u["login_id"], "name": u["name"], "designation": u["designation"],
                    "school_id": u.get("school_id"), "passcode": u.get("passcode")} for u in demo],
}
json.dump(meta, open(os.path.join(OUT, "meta.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=2)

def sql(v): return "NULL" if v is None or v == "" else "'" + str(v).replace("'", "''") + "'"
lines = ["-- Auto-generated by scripts/generateSeed.py", "begin;", ""]
for e in entities:
    lines.append(f"insert into entities (id, name, name_gu, level, parent_id, meta) values ({sql(e['id'])}, {sql(e['name'])}, {sql(e['name_gu'])}, {sql(e['level'])}, {sql(e['parent_id'])}, {sql(json.dumps(e['meta'], ensure_ascii=False))}::jsonb);")
lines.append("")
for u in app_users:
    lines.append(f"insert into app_users (id, login_id, name, name_gu, role, designation, entity_id, school_id, passcode, active) values ({sql(u['id'])}, {sql(u['login_id'])}, {sql(u['name'])}, {sql(u.get('name_gu'))}, {sql(u['role'])}, {sql(u['designation'])}, {sql(u['entity_id'])}, {sql(u.get('school_id'))}, {sql(u.get('passcode'))}, {str(u['active']).lower()});")
lines += ["", "commit;"]
open(os.path.join(SUPA, "seed.sql"), "w", encoding="utf-8").write("\n".join(lines))

print("Seed written:", counts)
print("Districts:", ", ".join(meta["districts"]), "· PM SHRI:", meta["pmShriSchools"], "· principals anon:", meta["principalsAnonymised"])
for u in demo:
    print(f"  {u['role']:<10} {str(u['login_id']):<12} {'PIN '+u['passcode'] if u.get('passcode') else 'UDISE '+str(u.get('school_id'))}")
